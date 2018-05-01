import scrapy
import datetime
import pprint
import openpyxl
from calendar import Calendar
from pprint import pprint
from data_ulis import get_list_of_N_day_ago
from data_ulis import get_color
def get_urls(years,months):# NOTE: Don't use anymore
    start_urls=[]
    href='http://ketqua.net/xo-so-mien-bac.php?ngay='
    now=datetime.datetime.now()
    current_date= now.date()
    current_time = now.time()
    today6pm = now.replace(hour=18, minute=30, second=0, microsecond=0)
    for year in years:
        for mo in months:
            days = Calendar().itermonthdates(year,mo)
            for day in days:
                if day.month == mo:

                    if day<current_date:
                        url=href+ day.strftime('%d-%m-%Y')
                        start_urls.append(url)
                    elif day==current_date:
                        if now > today6pm:
                            url=href+ day.strftime('%d-%m-%Y')
                            start_urls.append(url)
                            break
                        else:
                            #update date to write initial col
                            break

    return start_urls


class lottery(scrapy.Spider):
    name='lot2'

    def __init__(self):
        self.file_path='./excels/lottery-2007.xlsx'
        try:
            wb=openpyxl.load_workbook(self.file_path)
            self.exist=True
        except:
            wb=openpyxl.Workbook()
            wb.save(self.file_path)
            self.exist=False


        #get the max row to update
        wb=openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        self.row=sheet.max_row if sheet.max_row !=0 else 1#keep track current line to write
        self.date = ''


        #generate urls

        # days,self.start_urls=get_list_of_N_day_ago(5970)
        self.start_urls=get_urls([2007],[1,2,3,4,5,6,7,8,9,10,11,12])
        self.current=0#keep track current date to write


    #create freeze line
    def free_panes(self):

        wb=openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        for i in range(0,100):
            sheet.cell(row=self.row,column=i+2).value = i
        sheet.freeze_panes='A2'
        self.row+=1
        wb.save(self.file_path)

    def start_requests(self):
        if not self.exist:
            self.free_panes()

        for url in self.start_urls:
            date= url.split('=')
            yield scrapy.Request(url=url, callback=self.parse,meta={'date':date[1]})



    def parse(self,response):
        divs = response.css('.chu17.need_blank::text')
        ###date don't have number will not save
        if len(divs) >10:
            numbers = dict((el,0) for el in range(100))
            for div in divs:

                try:
                    number = int(div.extract())
                    numbers[number]+=1
                except:
                    pass

            wb=openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            #write datetime
            sheet.cell(row=self.row,column=1).value=response.meta.get('date')


            ##write numbers
            for idx,val in numbers.items():
                if val!=0:
                    sheet.cell(row=self.row,column=idx+2).value=val
                    sheet.cell(row=self.row,column=idx+2).fill=get_color(val)
            self.row+=1

            wb.save(self.file_path)

        #update current line
        else:
            print('=========================skipped :',response.meta.get('date'))
