import scrapy
import datetime
import pprint
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from calendar import Calendar
from pprint import pprint
from data_ulis import get_list_of_N_day_ago



def get_color(number):
    if number ==1 :
        return PatternFill(start_color='2ecc71',end_color='2ecc71',fill_type='solid')#green
    elif number ==2:
        return  PatternFill(start_color='2980b9',end_color='2980b9',fill_type='solid')#blue
    elif number==3:
        return  PatternFill(start_color='f1c40f',end_color='f1c40f',fill_type='solid')#yellow
    elif number ==4:
        return  PatternFill(start_color='e74c3c',end_color='e74c3c',fill_type='solid')#red
    else:
        return  PatternFill(start_color='FFFF0000',end_color='FFFFFF',fill_type='solid')#white



class lottery(scrapy.Spider):
    name='lot'

    def __init__(self):
        self.file_path='./excels/lottery100.xlsx'
        try:
            wb=openpyxl.load_workbook(self.file_path)
            self.exist=True
        except:
            wb=openpyxl.Workbook()
            wb.save(self.file_path)
            self.exist=False


        #get the max row to update
        wb=openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        self.row=sheet.max_row if sheet.max_row !=0 else 1#keep track current line to write
        self.date = ''


        #generate urls

        days,self.start_urls=get_list_of_N_day_ago(100)
        self.current=0#keep track current date to write

    #create freeze line
    def free_panes(self):

        wb=openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        for i in range(0,100):
            sheet.cell(row=self.row,column=i+2).value = i
        sheet.freeze_panes='A2'
        self.row+=1
        wb.save(self.file_path)



    def start_requests(self):
        if not self.exist:
            self.free_panes()

        date= self.start_urls[self.current].split('=')
        yield scrapy.Request(url=self.start_urls[self.current],callback=self.parse,meta={'date':date[1]})



    def parse(self,response):
        divs = response.css('.chu17.need_blank::text')
        ###date don't have number will not save
        if len(divs)!=0:
            numbers = dict((el,0) for el in range(100))
            for div in divs:
                number = int(div.extract())
                numbers[number]+=1


            wb=openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            #write datetime
            sheet.cell(row=self.row,column=1).value=response.meta.get('date')


            ##write numbers
            for idx,val in numbers.items():
                if val!=0:
                    sheet.cell(row=self.row,column=idx+2).value=val
                    sheet.cell(row=self.row,column=idx+2).fill=get_color(val)
            self.row+=1

            wb.save(self.file_path)

        #update current line
        self.current+=1
        date= self.start_urls[self.current].split('=')
        yield scrapy.Request(url=self.start_urls[self.current],callback=self.parse,meta={'date':date[1]})
