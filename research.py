import numpy as np
import openpyxl
from data_ulis import update
from collections import Counter
from openpyxl.styles import Color, PatternFill
file_path = './excels/lottery.xlsx'
save_path = './excels/analysis.xlsx'


def get_percentage(sheet):
    ###compute percentage
    max_row=sheet.max_row+2
    max_col=sheet.max_column

    row=max_row+1
    sheet.cell(row=row,column=1).value='Tỉ lệ %'
    for i in range(2,max_col+1):
        sheet.cell(row=row,column=i).value='=SUM(INDIRECT(ADDRESS(1,COLUMN())&":"&ADDRESS(ROW()-1,COLUMN())))*100/'+str(max_row)

def free_panes(sheet,position):
    # for i in range(0,100):
    #     sheet.cell(row=1,column=i+2).value = i
    sheet.freeze_panes=position

def get_color(number):


    if number in range(0,6):
        return PatternFill(start_color='95a5a6',end_color='95a5a6',fill_type='solid')#gray

    elif number in range(6,11):
        return PatternFill(start_color='2ecc71',end_color='2ecc71',fill_type='solid')#green

    elif number in range(11,16):
        return  PatternFill(start_color='2980b9',end_color='2980b9',fill_type='solid')#blue

    elif number in range(16,20):
        return  PatternFill(start_color='f1c40f',end_color='f1c40f',fill_type='solid')#yellow
    elif number >=20:
        return  PatternFill(start_color='e74c3c',end_color='e74c3c',fill_type='solid')#red



def get_frequency(sheet):
    max_row=sheet.max_row+1
    max_col=sheet.max_column

    row=max_row +1
    sheet.cell(row=row,column=1).value='Chuỗi:'
    row=max_row +1 +1
    sheet.cell(row=row,column=1).value='Trung bình xuất hiện:'
    row=max_row +1 +1 +1
    sheet.cell(row=row,column=1).value='Chuỗi dài nhất:'
    row=max_row +1 +1 +1 +1
    sheet.cell(row=row,column=1).value=':Lọc chuỗi:'

    for i in range(1,50):#print 50 occurencies
        row = max_row +1 +1 +1 +1 + i
        sheet.cell(row=row,column=1).value=i


    for i in range(2,max_col+1):#loop through column
        frequency=0
        appeareance=[]
        result = ''
        max=0
        for l in range(1,max_row+1+1):#loop through rows
            if sheet.cell(row=l,column=i).value is not None:#when a number appear

                if frequency==0:
                    result+='*'

                else:
                    result+=str(frequency)+'*'

                appeareance.append(frequency)
                frequency=0
                max=np.max(appeareance)

            else:
                frequency+=1
        sheet.cell(row=l,column=i).value=result#+1 because don't overwrite percentage
        sheet.cell(row=l+1,column=i).value=np.mean(appeareance)
        sheet.cell(row=l+2,column=i).value=max





        for key,val in Counter(appeareance).items():
            if key == 0:
                continue
            sheet.cell(row=l+2 +1 + key ,column=i).fill=get_color(val)
            sheet.cell(row=l+2 +1 + key ,column=i).value=val






def main():
    update('excels/lottery.xlsx')

    file_path='excels/lottery.xlsx'
    wb=openpyxl.load_workbook(file_path)

    for sheet_name in ['100','300','800','all']:
        sheet=wb[sheet_name]
        sheet.freeze_panes=sheet['B2']
        get_percentage(sheet)
        get_frequency(sheet)



    wb.save(save_path)










if __name__=='__main__':
    main()
